import pandas as pd
from openpyxl import load_workbook
import datetime
import os
import locale
import time
import PySimpleGUI as psg
from openpyxl.styles import PatternFill

class tratamento:
    def __Init__(self):
        pass
    def DRE(self,arquivo):
        # Iniciar o temporizador
        start_time = time.time()

        # Configurar a localização para o Brasil (pt_BR)
        locale.setlocale(locale.LC_ALL, 'pt_BR')

        # Obter a data atual
        data_atual = datetime.date.today()
        ano = data_atual.year

        '''inicio da tratativa do dataframe
        --------------------------------------------------------------------------------------------'''
        # Lê o arquivo 'dre.xlsx' e armazena os dados em um DataFrame df.
        #df = pd.read_excel(os.path.abspath('dre.xlsx'))
        df = pd.read_excel(arquivo)
        
        movimentacao = df.iloc[3,0]
        coluna_movimentacao = movimentacao.split(' / ')
        coluna_movimentacao = coluna_movimentacao[1].split(': ')[1]
        
        unidade = df.iloc[4,0]
        coluna_unidade = unidade.split(' / ')
        coluna_unidade = coluna_unidade[0]

        # Remover as primeiras 8 linhas
        df = df.drop(range(0, 8))
        # Redefinir o índice
        df = df.reset_index(drop=True)

        # Concatenar as duas primeiras linhas para formar o novo cabeçalho
        primeira_linha = df.iloc[0].astype(str)
        segunda_linha = df.iloc[1].astype(str)
        novo_cabecalho = primeira_linha + '-' + segunda_linha
        #define a primeira linha como cabeçado e remove os espaço dela
        novo_cabecalho = novo_cabecalho.str.replace(' ', '').str.replace('-', ' ').str.replace('nan', '').str.replace('ç', 'c').str.replace('ã', 'a')
        df.columns = novo_cabecalho

        '''--Lista de meses--'''
        # Extrair apenas os meses usando uma expressão regular
        segunda_linha_meses = segunda_linha.str.extract(r'(\w{3})/\d{4}').squeeze()
        # Remover valores duplicados e valores 'NaN' da série e transformar em uma lista
        meses = segunda_linha_meses.drop_duplicates().dropna().tolist()

        # Filtrar colunas que contêm a substring 'Saldo', 'Débitos', 'Créditos', 'Metas/Orçam.' e '%Mt/Or'
        colunas_a_remover = df.columns[df.columns.str.contains('Saldo|Débitos|Créditos|Metas/Orcam.|%Mt/Or')]
        df = df.drop(columns=colunas_a_remover)

        # Redefinir o índice
        df = df.reset_index(drop=True)
        df = df.drop(range(0, 2))

        #procurando valores CR,DB e removendo de cada celula
        def clean_value(x):
            if isinstance(x, str) and 'DB' in x:
                x = x.replace('DB', '').strip().replace(".", "").replace(",", ".")
                x = "-" + x
                #x = float(x) * -1
                return x
            elif isinstance(x, str) and 'CR' in x:
                x = x.replace('CR', '').strip().replace(".", "").replace(",", ".")
                #x = float(x)
                return x
            return x

        df.iloc[:, 4:] = df.iloc[:, 4:].map(clean_value)

        '''Inicio da procura e calculo necessario
                    --------------------------------------------------------------------------------------------'''
        #dicionario a ser preenchido
        #Planilha DRE
        contas_resultado = {}
        provisao_de_imposto = {}
        receitas_operacaional = {}
        custo_mercadorias = {}
        receitas_operacaional = {}
        venda_mercadorias = {}
        recuperacao_despesas = {}
        contribuicao_social = {}
        recuperacao_despesa = {}
        
        # Resumo    
        resultado_liquido = {}
        margem_contabil = {}
        lucro = {}
        receitas_nao_operacionais = {}
        recolhido = {}
        lucro_contabil = {}
        venda = {}
        ir = {}
        add = {}
        csll = {}
        total = {}
        total_ir_add = {}

        #Trimestre
        tri1 = {'RESULTADO LIQUIDO': 0,'Margem Contábil': 0,'Lucro': 0,'Receitas Não Operacionais': 0,'Recolhido': 0,
        'Lucro Contábil': 0,'Venda': 0,'IR': 0,'Add': 0,'Csll': 0,'Total': 0,'Total IR-Add': 0}
        tri2 = {'RESULTADO LIQUIDO': 0,'Margem Contábil': 0,'Lucro': 0,'Receitas Não Operacionais': 0,'Recolhido': 0,
        'Lucro Contábil': 0,'Venda': 0,'IR': 0,'Add': 0,'Csll': 0,'Total': 0,'Total IR-Add': 0}
        tri3 = {'RESULTADO LIQUIDO': 0,'Margem Contábil': 0,'Lucro': 0,'Receitas Não Operacionais': 0,'Recolhido': 0,
        'Lucro Contábil': 0,'Venda': 0,'IR': 0,'Add': 0,'Csll': 0,'Total': 0,'Total IR-Add': 0}
        tri4 = {'RESULTADO LIQUIDO': 0,'Margem Contábil': 0,'Lucro': 0,'Receitas Não Operacionais': 0,'Recolhido': 0,
        'Lucro Contábil': 0,'Venda': 0,'IR': 0,'Add': 0,'Csll': 0,'Total': 0,'Total IR-Add': 0}

        def valor_da_linha(linha_descricao, coluna_a_encontrar,contem=True):
            coluna_filtro = 'DescricaoConta '
            try:
                # Tratar valores ausentes substituindo por uma string vazia
                df[coluna_filtro] = df[coluna_filtro].fillna('')
                if contem:
                    #encontrar linha a partir de uma coluna como filtro e contendo os valores
                    linha_contas_resultado = df[df[coluna_filtro].str.contains(linha_descricao, case=False)]
                else:
                    # Localizar a célula que contém o texto inserido
                    linha_contas_resultado = df[df[coluna_filtro].str.strip() == linha_descricao]
                # Obter o valor da coluna inserido
                valor_linha = linha_contas_resultado[coluna_a_encontrar].values[0]
                indice_linha = linha_contas_resultado.index[linha_contas_resultado[coluna_a_encontrar] == valor_linha].tolist()[0]
                num_coluna = df.columns.get_loc(coluna_a_encontrar)
                valor_linha = linha_contas_resultado[coluna_a_encontrar].str.strip().values[0]
                if valor_linha == '':
                    valor_linha = 0
                else:
                    valor_linha = float(linha_contas_resultado[coluna_a_encontrar].values[0])
                return valor_linha, indice_linha , num_coluna
            except Exception as e:
                print(e)
                return 0
            
        for mes in meses:
            #definindo o ano e mes para procurar em relação a função mês
            coluna_a_encontrar = f'MvtoLíquido {mes}/{ano}'
            
            def preencher_dicionario(linha, dicionario, contem_linha=True):
                valor_linha, indice_linha, num_coluna = valor_da_linha(linha, coluna_a_encontrar, contem_linha)
                dicionario[f'MvtoLíquido {mes}/{ano}'] = valor_linha
                dicionario[f'Index_linha {mes}/{ano}'] = indice_linha
                dicionario[f'Index_coluna {mes}/{ano}'] = num_coluna
                
            #atualizando o dicionario
            def att(dataset,resultado):
                for item in dataset.items():
                    indice1_dicionario = item[:2]
                    if indice1_dicionario[1] == dataset[coluna_a_encontrar]: 
                        dataset.update({ item[0]: resultado })
                        break
                        
            preencher_dicionario('CONTAS DE RESULTADO',contas_resultado)
            preencher_dicionario('PROVISAO DE IMPOSTO S/L', provisao_de_imposto)
            preencher_dicionario('RECEITAS OPERACIONAL LIQUIDA', receitas_operacaional)
            preencher_dicionario('CUSTO DAS MERCADORIAS VENDIDOS - CMV', custo_mercadorias)
            preencher_dicionario('RECEITAS OPERACIONAL LIQUIDA', receitas_operacaional)
            preencher_dicionario('VENDA DE MERCADORIAS',venda_mercadorias)
            preencher_dicionario('Recuperacao De Despesas Exerc Anterior',recuperacao_despesas)
            preencher_dicionario('Contribuicao Social',contribuicao_social,False)
            preencher_dicionario('Recuperacao De Despesas Exerc Anterior', recuperacao_despesa)
            
            #contas de resultado calculo                 
            resultado_analisado = contas_resultado[coluna_a_encontrar] + provisao_de_imposto[coluna_a_encontrar] * - 1
            att(contas_resultado,resultado_analisado)


        # Salvar planilha Excel tratada, trocando index = True mostra o index das linhas
        df.to_excel('dre_tratada.xlsx', index=False)

        '''definindo a primeira linha como cabeçalho, copiando o dataframe e criando novas aba com os nomes (Resumo, Planilha DRE)
                    --------------------------------------------------------------------------------------------'''
        # Carregar o arquivo Excel tratado
        wb = load_workbook('dre_tratada.xlsx')
        # Renomear a planilha 'Sheet1' para 'Planilha DRE' 
        if 'Sheet1' in wb.sheetnames:
            sheet = wb['Sheet1']
            sheet.title = 'Planilha DRE'
        # Cria uma nova planilha antes da planilha ativa
        wb.create_sheet('Resumo', 0)
        # Salvar novamente o arquivo
        wb.save('dre_tratada.xlsx')

        # Lê o arquivo 'dre_tratada.xlsx' e armazena os dados em um DataFrame df_sheet1
        df_sheet1 = pd.read_excel(os.path.abspath('dre_tratada.xlsx'), sheet_name='Planilha DRE')
        # Lê o arquivo 'dre_tratada.xlsx' e armazena os dados em um DataFrame df_resumo
        df_resumo = pd.read_excel(os.path.abspath('dre_tratada.xlsx'), sheet_name='Resumo')
        df_resumo.ffill(inplace=True)

        # Definindo os cabeçalhos resumo
        cabecalhos = pd.Series(segunda_linha[segunda_linha != 'nan'].str.strip().drop_duplicates())
        cabecalhos['Unnamed: 2'] = 'DescriçãoConta '
        novo_cab_res = {}
       
        for k, v in cabecalhos.items():
            if v == 'Mar/2023':
                novo_cab_res[k] = v
                novo_cab_res['Unnamed: 18'] = '1 Tri'    
            elif v == 'Jun/2023':
                novo_cab_res[k] = v
                novo_cab_res['Unnamed: 36'] = '2 Tri'
            elif v == 'Set/2023':
                novo_cab_res[k] = v
                novo_cab_res['Unnamed: 46'] = '3 Tri'
            elif v == 'Dez/2023':
                novo_cab_res[k] = v
                novo_cab_res['Unnamed: 56'] = '4 Tri'
                break
            novo_cab_res[k] = v

        # Preenchendo os cabeçalhos na primeira linha do DataFrame
        for col, cabecalho in novo_cab_res.items():
            df_resumo.at[1, col] = cabecalho

        # Definir a primeira linha como cabeçalho
        novo_cabecalho = df_resumo.iloc[0].astype(str)
        df_resumo.columns = novo_cabecalho
        # Remover a primeira linha, que agora é o cabeçalho duplicado
        df_resumo = df_resumo.iloc[1:].reset_index(drop=True)

        linhas = ['Venda','Lucro Contábil','Recolhido','Receitas Não Operacionais','Lucro','-','IR','Add' ,'Csll','Total','-', 'Total IR-Add','-', 'Margem Contábil','RESULTADO LIQUIDO']

        for lin, linha in enumerate(linhas):
            df_resumo.at[lin + 2, 'DescriçãoConta '] = linha

        def valor_da_linha_resumo(linha_descricao, coluna_a_encontrar,contem=True):
            coluna_filtro = 'DescriçãoConta '
            try:
                # Tratar valores ausentes substituindo por uma string vazia
                df_resumo[coluna_filtro] = df_resumo[coluna_filtro].fillna('')
                if contem:
                    #encontrar linha a partir de uma coluna como filtro e contendo os valores
                    linha_contas_resultado = df_resumo[df_resumo[coluna_filtro].str.contains(linha_descricao, case=False)]
                else:
                    # Localizar a célula que contém o texto inserido
                    linha_contas_resultado = df_resumo[df_resumo[coluna_filtro].str.strip() == linha_descricao]
                # Obter o valor da coluna inserido
                valor_linha = linha_contas_resultado[coluna_a_encontrar].values[0]
                indice_linha = linha_contas_resultado.index[0]
              
                num_coluna = df_resumo.columns.get_loc(coluna_a_encontrar)
                
                if type(valor_linha) == str:
                    valor_linha = 0
                return valor_linha, indice_linha , num_coluna
            except Exception as e:
                print(e)
                return 0
            
        for k, mes in enumerate(meses):
            coluna_a_encontrar = f'{mes}/{ano}'
            def preencher_dicionario_resumo(linha, dicionario, contem_linha=True):
                venda, index_linha, index_coluna = valor_da_linha_resumo(linha, coluna_a_encontrar,contem_linha)
                dicionario[f'MvtoLíquido {mes}/{ano}'] = venda
                dicionario[f'Index_linha {mes}/{ano}'] = index_linha - 2
                dicionario[f'Index_coluna {mes}/{ano}'] = index_coluna
            
            def preencher_valor_planilha(linha,calculo):
                linha_ = linha[f'Index_linha {mes}/{ano}']
                coluna_ = linha[f'Index_coluna {mes}/{ano}']
                df_resumo.iloc[linha_,coluna_] = calculo
                
            preencher_dicionario_resumo('RESULTADO LIQUIDO',resultado_liquido)
            preencher_dicionario_resumo('Margem Contábil',margem_contabil)
            preencher_dicionario_resumo('Lucro',lucro,False)
            preencher_dicionario_resumo('Receitas Não Operacionais',receitas_nao_operacionais)
            preencher_dicionario_resumo('Recolhido',recolhido)
            preencher_dicionario_resumo('Lucro Contábil',lucro_contabil)
            preencher_dicionario_resumo('Venda',venda)
            preencher_dicionario_resumo('IR',ir)
            preencher_dicionario_resumo('Add',add)
            preencher_dicionario_resumo('Csll',csll)
            preencher_dicionario_resumo('Total',total)
            preencher_dicionario_resumo('Total IR-Add',total_ir_add)

            def calculo(mes1):
                def cond_tri(atributo, res):
                    if mes1 == f'MvtoLíquido Jan/{ano}' or mes1 == f'MvtoLíquido Fev/{ano}' or mes1 == f'MvtoLíquido Mar/{ano}':
                        va = tri1[atributo]
                        tri1[atributo] = va + res
                    elif mes1 == f'MvtoLíquido Abr/{ano}' or mes1 == f'MvtoLíquido Mai/{ano}' or mes1 == f'MvtoLíquido Jun/{ano}':
                        va = tri2[atributo]
                        tri2[atributo] = va + res
                    elif mes1 == f'MvtoLíquido Jul/{ano}' or mes1 == f'MvtoLíquido Ago/{ano}' or mes1 == f'MvtoLíquido Set/{ano}':
                        va = tri3[atributo]
                        tri3[atributo] = va + res
                    else:
                        va = tri4[atributo]
                        tri4[atributo] = va + res
                try:
                    #calculo e preenchimento da venda
                    calculo_venda_1 =venda_mercadorias[mes1]
                    calculo_venda_1 = round(calculo_venda_1,2)
                    cond_tri('Venda',calculo_venda_1)
                    preencher_valor_planilha(venda, calculo_venda_1)

                    #calculo e preenchimento lucro contabil
                    lucro_c = contas_resultado[mes1]
                    lucro_c = round(lucro_c,2)
                    cond_tri('Lucro Contábil',lucro_c)
                    preencher_valor_planilha(lucro_contabil, lucro_c)

                    #calculo e preenchimento Receitas Não Operacionais
                    calculo_receita =recuperacao_despesas[mes1]
                    calculo_receita = round(calculo_receita,2)
                    cond_tri('Receitas Não Operacionais',calculo_receita)
                    preencher_valor_planilha(receitas_nao_operacionais, calculo_receita)

                    #calculo e preenchimento lucro
                    reultado_lucro = lucro_c - calculo_receita
                    reultado_lucro = round(reultado_lucro,2)
                    cond_tri('Lucro',reultado_lucro)
                    preencher_valor_planilha(lucro,reultado_lucro)
                    
                    #calculo e preenchimento recolhido
                    cont_social = contribuicao_social[f'MvtoLíquido {mes}/{ano}']
                    calculo = cont_social / 9 * 100
                    calculo = round(calculo,2)
                    cond_tri('Recolhido',calculo)
                    preencher_valor_planilha(recolhido,calculo * -1)

                    #calculo e preenchimento imposto de renda
                    calculo_IR = lucro_c * 0.15
                    calculo_IR = round(calculo_IR,2)
                    cond_tri('IR',calculo_IR)
                    preencher_valor_planilha(ir,calculo_IR)

                    #calculo e preenchimento add
                    calculo_add = (lucro_c - 60000) * 0.1
                    calculo_add = round(calculo_add,2)
                    cond_tri('Add',calculo_add)
                    preencher_valor_planilha(add, calculo_add)

                    #calculo e preenchimento csll
                    calculo_csll = lucro_c * 0.9 / 10
                    calculo_csll = round(calculo_csll,2)
                    cond_tri('Csll',calculo_csll)
                    preencher_valor_planilha(csll, calculo_csll)

                    #calculo e preenchimento total
                    calculo_total = calculo_IR + calculo_add + calculo_csll
                    calculo_total = round(calculo_total,2)
                    cond_tri('Total',calculo_total)
                    preencher_valor_planilha(total, calculo_total)
                    
                    #calculo e preenchimento total-ir-add
                    calculo_total_ir_add = calculo_IR + calculo_add
                    cond_tri('Total IR-Add',calculo_total_ir_add)
                    preencher_valor_planilha(total_ir_add, calculo_total_ir_add)
                except Exception as e:
                    print(e)

            #preenchimento resumo
            calculo(f'MvtoLíquido {mes}/{ano}')

            for i in range(1,4):
                trimestre = f'{i} Tri'
                mes_tri = ' '.join(i for i in novo_cab_res.values())
                def calc_tri(col,biblioteca):
                    coluna = col
                    #venda
                    linha = venda[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['Venda']
                    #lucro contabil
                    linha = lucro_contabil[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['Lucro Contábil']
                    #Receitas Não Operacionais
                    linha = receitas_nao_operacionais[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['Receitas Não Operacionais']
                    #Lucro
                    linha = lucro[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['Lucro']
                    #Recolhido
                    linha = recolhido[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['Recolhido']
                    #imposto de renda
                    linha = ir[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['IR']
                    #add
                    linha = add[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['Add']
                    #csll
                    linha = csll[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['Csll']
                    #total
                    linha = total[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['Total']
                    #total_ir_add
                    linha = total_ir_add[f'Index_linha {mes}/{ano}']
                    df_resumo.iloc[linha,coluna] = biblioteca['Total IR-Add']
                    
                if trimestre in mes_tri:
                    if trimestre == '1 Tri':
                        calc_tri(4,tri1)
                    if trimestre == '2 Tri':
                        calc_tri(8,tri2)
                    if trimestre == '3 Tri':
                        calc_tri(12,tri3)
                    if trimestre == '4 Tri':
                        calc_tri(16,tri4)
                        
            #calculo e preenchimento do resultado liquido
            valor_resultado1 = contas_resultado[f'MvtoLíquido {mes}/{ano}']
            valor_resultado2 = recuperacao_despesa[f'MvtoLíquido {mes}/{ano}']
            
            if valor_resultado2 != 0:
                calculo_resultado = valor_resultado1 - valor_resultado2
                preencher_valor_planilha(resultado_liquido,calculo_resultado)
            else:    
                preencher_valor_planilha(resultado_liquido,  contas_resultado[f'MvtoLíquido {mes}/{ano}'])

            #calculo e preenchimento da margem contabil
            valor_contabil1 = receitas_operacaional[f'MvtoLíquido {mes}/{ano}']
            valor_contabil2 = custo_mercadorias[f'MvtoLíquido {mes}/{ano}']
            
            if valor_contabil2 != 0:
                calculo_contabil = valor_contabil1 + valor_contabil2
                preencher_valor_planilha(margem_contabil,calculo_contabil)
            else:    
                preencher_valor_planilha(margem_contabil, contas_resultado[f'MvtoLíquido {mes}/{ano}'])
                            
        # Salvar planilhas Excel tratadas, trocando index = True mostra o index das linhas
        with pd.ExcelWriter(f'dre_tratada.xlsx') as writer:
            df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
            df_sheet1.to_excel(writer, sheet_name='Planilha DRE', index=False)

        # Finalizar o temporizador
        end_time = time.time()
        # Calcular o tempo de execução
        elapsed_time = end_time - start_time

        print(f"Tempo total de execução: {elapsed_time:.2f} segundos")

        # Carregar o arquivo Excel tratado
        wb = load_workbook('dre_tratada.xlsx')
        # Especifique o nome da aba que deseja modificar
        nome_da_aba = 'Planilha DRE'
        # Acesse a aba desejada
        aba_existente = wb[nome_da_aba]

        def cor_fundo_celula(numero_index_linha, cor_hex):
            # Especifique a cor de fundo que deseja aplicar #ffffff
            cor_de_fundo = cor_hex 
            # Percorra todas as células da linha especificada e defina a cor de fundo
            for cell in aba_existente[numero_index_linha]:
                cell.fill = PatternFill(start_color=cor_de_fundo, end_color=cor_de_fundo, fill_type='solid')
            
        def encontrar_indices_texto(linha_descricao):
        # Preencher os valores ausentes na coluna de interesse com um valor padrão
            df_filled = df.fillna({'DescricaoConta ': ''})
            # Localizar as células que contêm o texto inserido
            linhas_contas_resultado = df_filled[df_filled['DescricaoConta '].str.contains(pat=linha_descricao)]
            return linhas_contas_resultado.index.tolist()

        lista_descrição_azul = (
        'CONTAS DE RESULTADO', 'RECEITAS', 'RECEITAS OPERACIONAL LIQUIDA', 'DEDUCOES DA VENDA DE MERCADORIAS','VENDA DE MERCADORIAS' ,
        'RECEITAS COMERCIAIS', 'RECEITAS DIVERSAS', 'RECEITAS FINANCEIRAS', 'RECEITAS PRESTACOES SERVICOS',
        'CUSTO DOS SERVICOS PRESTADOS', 'CUSTO DOS SERVIÇOS PRESTADOS' , 'CUSTOS DAS MERCADORIAS VENDIDOS', 'CUSTO DAS MERCADORIAS VENDIDOS - CMV',
        'ESTOQUE INICIAL', 'ENTRADAS DE MERCADORIAS', 'ESTOQUE FINAL', 'CUSTO OPERACIONAL', 'CUSTO DE PRODUCAO',
        'DESPESAS OPERACIONAIS', 'DESPESAS COM PESSOAL', 'PRO LABORE', 'DESPESAS COM PESSOAL DIRETO', 'BENEFICIOS',
        'REMUNERACAO VARIAVEIS', 'SERVICOS TERCEIRIZADOS', 'OUTRAS DESPESAS OPERACIONAIS', 'MANUTENCAO', 'SEGUROS',
        'DESPESAS COM VEICULOS', 'TRANSPORTES E DESLOCAMENTOS', 'VIAGENS E ESTADIAS', 'ENERGIA', 'COMUNICACAO',
        'OUTRAS DESPESAS VENDAS', 'DESPESAS EMBALAGENS', 'DESPESAS CONSUMO INTERNO', 'MANUTENCAO DE IMOVEIS',
        'DESPESAS COM INFORMATICA', 'DESPESAS GERAIS', 'MARKETING', 'DESPESAS DE EXPEDIENTE', 'ALUGUEIS',
        'IMPOSTOS E TAXAS', 'DEPRECIACOES E PROVISOES', 'DESPESAS DIVERSAS', 'HONORARIOS PROFISSIONAIS',
        'DESPESAS FINANCEIRAS', 'SUPERMERCADO AQUINO', 'ASSOCIACOES DE CLASSE', 'DESPESAS FISCAIS',
        'RESULTADO DO EXERCICIO', 'DESPESAS E RECEITAS', 'RECEITAS OPERACIONAIS', 'DESPESAS OPERACIONAIS',
        'CUSTO DA MERCADORIAS VENDIDAS', 'RECEITAS E DESPESAS NAO OPERACIONAIS', 'RESULTADO NA VENDA IMOBILIZADO',
        'PROVISAO DE IMPOSTO S/L', 'PROVISAO DE IMPOSTO S/L', 'TRANSF. PARA RESERVA DE L', 'TRANFERENCIA PARA RESERVA DE L'
    )

        # faz um loop for com a lista para pintar as linhas cuja descrição tem na lista
        for descricao in lista_descrição_azul:
            indices = encontrar_indices_texto(descricao)
            for index in indices:
                index_int = int(index)  
                cor_fundo_celula(index_int, 'C5D9F1')

        cor_fundo_celula(4, 'FFFF00') #amarelo

        # Salvar novamente o arquivo
        wb.save('dre_tratada.xlsx')

if __name__ == "__main__":
    tratamento = tratamento()
    selected_theme = 'Reddit'
    psg.theme(selected_theme)
    file = psg.popup_get_file('Selecione o arquivo bruto',  title="Tratamento DRE", keep_on_top=True,icon=os.path.abspath('iconp.ico'))

    if file == None:
        psg.popup_error('Arquivo não selecionado!', title=" ", keep_on_top=True,icon=os.path.abspath('iconp.ico'))
    else:
        tratamento.DRE(file)
    
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
import os
import re
import locale

class programa:
    def __init__(self):
        super().__init__()
        # Configurar a localização para o Brasil (pt_BR)
        locale.setlocale(locale.LC_ALL, 'pt_BR')

        UNIDADE = '001 - Supermercado Aquino.'
        MOVIMENTACAO = 'Movimentação Contábil Gerencial Mensal / Período Movimento: 01/01/23 a 30/06/23'

        # Obter a data atual
        data_atual = datetime.date.today()
        ano = data_atual.year

        try:
            '''inicio da tratativa do dataframe
            --------------------------------------------------------------------------------------------'''
            # Lê o arquivo 'dre.xlsx' e armazena os dados em um DataFrame df.
            df = pd.read_excel(os.path.abspath('dre.xlsx'))
            
            # Remover as primeiras 8 linhas
            df = df.drop(range(0, 8))
            # Redefinir o índice
            df = df.reset_index(drop=True)

            # Concatenar as duas primeiras linhas para formar o novo cabeçalho
            primeira_linha = df.iloc[0].astype(str)
            segunda_linha = df.iloc[1].astype(str)
            novo_cabecalho = primeira_linha + '-' + segunda_linha
            #define a primeira linha como cabeçado e remove os espaço dela
            df.columns = novo_cabecalho.str.replace(' ', '').str.replace('-', ' ').str.replace('nan', '')
            
            '''--Lista de meses--'''
            # Extrair apenas os meses usando uma expressão regular
            segunda_linha_meses = segunda_linha.str.extract(r'(\w{3})/\d{4}').squeeze()
            # Remover valores duplicados e valores 'NaN' da série e transformar em uma lista
            meses = segunda_linha_meses.drop_duplicates().dropna().tolist()
            
            # Filtrar colunas que contêm a substring 'Saldo', 'Débitos', 'Créditos', 'Metas/Orçam.' e '%Mt/Or'
            colunas_a_remover = df.columns[df.columns.str.contains('Saldo|Débitos|Créditos|Metas/Orçam.|%Mt/Or')]
            df = df.drop(columns=colunas_a_remover)
            
            print(df)
            #procurando valores CR e removendo de cada celula
            df = df.applymap(lambda x: x.replace('CR', '') if isinstance(x, str) else x)
            
            '''Inicio da procura e calculo necessario
            --------------------------------------------------------------------------------------------'''
            def encontrar_VALOR_positivo(coluna_filtro, linha_descricao, coluna_a_encontrar,contem=True):
                try:
                    # Tratar valores ausentes substituindo por uma string vazia
                    df[coluna_filtro] = df[coluna_filtro].fillna('')
                    if contem:
                        #encontrar linha a partir de uma coluna como filtro e contendo os valores
                        linha_contas_resultado = df[df[coluna_filtro].str.contains(linha_descricao, case=False)]
                    else:
                        # Localizar a célula que contém o texto inserido
                        linha_contas_resultado = df[df[coluna_filtro].str.strip() == linha_descricao]
                    # Obter o valor da coluna inserido
                    valor_linha = linha_contas_resultado[coluna_a_encontrar].str.replace(' ','').values[0]
                    return valor_linha
                except Exception as e:
                    return 0

            def encontrar_VALOR_negativo(coluna_filtro, linha_descricao,coluna_a_encontrar,contem=True):
                try:
                    # Tratar valores ausentes substituindo por uma string vazia
                    df[coluna_filtro] = df[coluna_filtro].fillna('')
                    if contem:
                        # Encontrar linha a partir de uma coluna como filtro e contendo os valores
                        linha_contas_resultado = df[df[coluna_filtro].str.contains(linha_descricao, case=False)]
                    else:
                        # Localizar a célula que contém o texto inserido
                        linha_contas_resultado = df[df[coluna_filtro].str.strip() == linha_descricao]
                    # Obter o índice da linha
                    indice_linha = linha_contas_resultado.index[0]
                    # Retornar o valor específico da coluna
                    valor_coluna = float(linha_contas_resultado.loc[indice_linha, coluna_a_encontrar].replace('DB', '').strip().replace(".", "").replace(",", "."))
                    # Retornar o índice da linha
                    return valor_coluna * -1
                except Exception as e:
                    return 0 
            
            def encontrar_VALOR_index(coluna_filtro, linha_descricao,coluna_a_encontrar,contem=True):
                try:
                    # Tratar valores ausentes substituindo por uma string vazia
                    df[coluna_filtro] = df[coluna_filtro].fillna('')
                    if contem:
                        # Encontrar linha a partir de uma coluna como filtro e contendo os valores
                        linha_contas_resultado = df[df[coluna_filtro].str.contains(linha_descricao, case=False)]
                    else:
                        # Localizar a célula que contém o texto inserido
                        linha_contas_resultado = df[df[coluna_filtro].str.strip() == linha_descricao]
                    # Obter o índice da linha
                    indice_linha = linha_contas_resultado.index[0]
                    # Obter o número da coluna onde o valor foi encontrado
                    num_coluna = df.columns.get_loc(coluna_a_encontrar)
                    # Retornar o índice da linha
                    return indice_linha, num_coluna 
                except Exception as e:
                    return None
                
            # Dicionário para armazenar os valores encontrados em cada mês de MvtoLíquido
            provisao_de_imposto = {}
            contas_resultado = {}
            recuperacao_despesa = {}
            receitas_operacaional = {}
            custo_mercadorias = {}
            venda_mercadorias = {}
            contribuicao_social = {}
            recuperacao_despesas = {}

            for mes in meses:
                #definindo o ano e mes para procurar em relação a função mês
                coluna_a_encontrar = f'MvtoLíquido {mes}/{ano}'
                coluna_descricao = 'DescriçãoConta '
                
                #função para encontrar o valor e a index da celula procurada
                def preencher_dicionario(linha,dicionario,contem_linha=True):
                    if contem_linha:
                        # Executando a função encontrar_VALOR_positivo e salvando em uma lista para calculo futuros    
                        encontrar_valor_positivo = encontrar_VALOR_positivo(coluna_descricao, linha, coluna_a_encontrar)
                        # Executando a função encontrar_VALOR_index para a localização da celula
                        encontrar_valor_index = encontrar_VALOR_index(coluna_descricao, linha,coluna_a_encontrar)
                    else:
                        # Executando a função encontrar_VALOR_positivo e salvando em uma lista para calculo futuros    
                        encontrar_valor_positivo = encontrar_VALOR_positivo(coluna_descricao, linha, coluna_a_encontrar,False)
                        # Executando a função encontrar_VALOR_index para a localização da celula
                        encontrar_valor_index = encontrar_VALOR_index(coluna_descricao, linha,coluna_a_encontrar,False)

                    if 'DB' in str(encontrar_valor_positivo):
                        if contem_linha:
                            # Executando a função encontrar_VALOR_negativo
                            encontrar_valor_negativo = encontrar_VALOR_negativo(coluna_descricao, linha,f'MvtoLíquido {mes}/{ano}')
                        else:
                            # Executando a função encontrar_VALOR_negativo
                            encontrar_valor_negativo = encontrar_VALOR_negativo(coluna_descricao, linha,f'MvtoLíquido {mes}/{ano}',False)

                        #Trocando a str DB por um float salvando na celula e na lista para calculos futuros
                        df.iloc[encontrar_valor_index[0], encontrar_valor_index[1]] = encontrar_valor_negativo
                        dicionario[f'MvtoLíquido_{mes.lower()}_{ano}'] = encontrar_valor_negativo
                        dicionario[f'Index_linha_{mes.lower()}_{ano}']  = encontrar_valor_index[0]
                        dicionario[f'Index_coluna_{mes.lower()}_{ano}']  = encontrar_valor_index[1]
                    else:
                        dicionario[f'MvtoLíquido_{mes.lower()}_{ano}'] = encontrar_valor_positivo
                        dicionario[f'Index_linha_{mes.lower()}_{ano}']  = encontrar_valor_index[0]
                        dicionario[f'Index_coluna_{mes.lower()}_{ano}']  = encontrar_valor_index[1]
            
                preencher_dicionario('CONTAS DE RESULTADO', contas_resultado)
                preencher_dicionario('PROVISAO DE IMPOSTO S/L', provisao_de_imposto)
                preencher_dicionario('Recuperacao De Despesas Exerc Anterior', recuperacao_despesa)
                preencher_dicionario('RECEITAS OPERACIONAL LIQUIDA', receitas_operacaional)
                preencher_dicionario('CUSTO DAS MERCADORIAS VENDIDOS - CMV', custo_mercadorias)
                preencher_dicionario('VENDA DE MERCADORIAS',venda_mercadorias)
                preencher_dicionario('Contribuicao Social',contribuicao_social,False)
                preencher_dicionario('Recuperacao De Despesas Exerc Anterior',recuperacao_despesas)

            # Dicionário para armazenar os valores encontrados em cada mês de MvtoLíquido sem DB
            def lista_sem_DB(lista):
                # Convertendo os valores do dicionário para strings e removendo os "DB"
                for key, value in lista.items():
                    value = re.sub(r'DB', '', str(value))
                    value = value.replace('.', '').replace(',', '.')
                    
                    # Verifica se a string está vazia
                    if not value:
                        # Define um valor padrão (0) para o caso da string estar vazia
                        lista[key] = 0
                    else:
                        try:
                            # Converte a string em float e atualiza o valor no dicionário
                            lista[key] = float(value)
                        except ValueError as e:
                            print(e)
                            # Caso ocorra um erro na conversão, trate-o aqui de acordo com sua lógica
                            # Por exemplo, definindo um valor padrão ou ignorando a entrada problemática
                            lista[key] = None
                            
                return lista
            
            lista_sem_DB(provisao_de_imposto)
            lista_sem_DB(contas_resultado)
            lista_sem_DB(recuperacao_despesa)
            lista_sem_DB(receitas_operacaional)
            lista_sem_DB(custo_mercadorias)
            lista_sem_DB(venda_mercadorias)
            lista_sem_DB(contribuicao_social)
            lista_sem_DB(recuperacao_despesas)

            #transfomando os meses todos em minusculo
            meses_minusculo = [mes.lower() for mes in meses]
            
            for mes in meses_minusculo:
                def att(dataset,resultado):
                    for item in dataset.items():
                        indice1_dicionario = item[:2]
                        if indice1_dicionario[1] == dataset[f'MvtoLíquido_{mes}_{ano}']: 
                            dataset.update({ item[0]: resultado })
                            break
                
                #idenficando se o numero é menor que 0
                if contas_resultado[f'MvtoLíquido_{mes}_{ano}'] < 0:
                    resultado_analisado = contas_resultado[f'MvtoLíquido_{mes}_{ano}'] + (provisao_de_imposto[f'MvtoLíquido_{mes}_{ano}'] * -1)
                    linha = int(contas_resultado[f'Index_linha_{mes}_{ano}'])
                    coluna = int(contas_resultado[f'Index_coluna_{mes}_{ano}'])
                    
                    # Formatando o resultado_analisado como string com o formato 'x.xxx.xxx,xx'
                    numero_formatado = "{:,.2f}".format(resultado_analisado / 100)
                    df.iloc[linha,coluna] =  numero_formatado
                    numero_formatado = float(numero_formatado.replace(',',''))
                    att(contas_resultado,numero_formatado)

                if custo_mercadorias[f'MvtoLíquido_{mes}_{ano}']:
                    resultado_analisado = custo_mercadorias[f'MvtoLíquido_{mes}_{ano}']
                    numero_formatado_custo = "{:,.2f}".format(resultado_analisado / 100)
                    numero_formatado_custo = float(numero_formatado_custo.replace(',',''))
                    att(custo_mercadorias,numero_formatado_custo)

            '''continuação do inicio da tratativa do dataframe
            --------------------------------------------------------------------------------------------'''
            #procurando valores DB e removendo de cada celula
            df = df.applymap(lambda x: x.replace('DB', '') if isinstance(x, str) else x)
            
            # Remover as primeiras 8 linhas
            df = df.drop(range(0, 2))
            # Redefinir o índice
            df = df.reset_index(drop=True)

            # Salvar planilha Excel tratada, trocando index = True mostra o index das linhas
            df.to_excel('dre_tratada.xlsx', index=False)

            '''criando uma nova aba - resumo
            --------------------------------------------------------------------------------------------'''
            # Carregar o arquivo Excel tratado
            wb = load_workbook('dre_tratada.xlsx')
            # Cria uma nova planilha antes da planilha ativa
            wb.create_sheet('Resumo', 0)
            # Salvar novamente o arquivo
            wb.save('dre_tratada.xlsx')

            '''definindo a primeira linha como cabeçalho, copiando os dataframe e criando novas aba com os nomes (Resumo, Planilha DRE)
            --------------------------------------------------------------------------------------------'''
            # Lê o arquivo 'dre_tratada.xlsx' e armazena os dados em um DataFrame df_sheet1
            df_sheet1 = pd.read_excel(os.path.abspath('dre_tratada.xlsx'), sheet_name='Sheet1')
            # Lê o arquivo 'dre_tratada.xlsx' e armazena os dados em um DataFrame df_resumo
            df_resumo = pd.read_excel(os.path.abspath('dre_tratada.xlsx'), sheet_name='Resumo')
            df_resumo.fillna(method='ffill', inplace=True)

            # Definindo os cabeçalhos resumo
            cabecalhos = pd.Series(segunda_linha[segunda_linha != 'nan'].str.strip().drop_duplicates())
            cabecalhos['Unnamed: 2'] = 'DescriçãoConta '
            

            #adicionar novos cabeçalhos
            '''def novos_cabecalho(cabecalhos, *args):
                for i, txt in enumerate(args, start=35):
                    cabecalhos[f'Unnamed: {i}'] = txt
            
            novos_cabecalho(cabecalhos, 'Base','IR','Add','Csll')'''

            # Preenchendo os cabeçalhos na primeira linha do DataFrame
            for col, cabecalho in enumerate(cabecalhos):
                df_resumo.at[1, col] = cabecalho
            
            # Adicionando a coluna vazia antes dos cabeçalhos
            df_resumo.insert(0, '', '')

            linhas = ['Venda','Lucro Contábil','Recolhido','Receitas Não Operacionais','Lucro','-','IR','Add' ,'Csll','Total', '-', 'Margem Contábil','RESULTADO LIQUIDO']
            
            for lin, linha in enumerate(linhas):
                df_resumo.at[lin + 2, 0] = linha
            
            # Preenchendo as informações da UNIDADE e Movimentação na terceira e quarta linha do DataFrame, respectivamente
            df_resumo.iloc[1, 0] = UNIDADE
            df_resumo.iloc[2, 0] = MOVIMENTACAO

            # Definir a primeira linha como cabeçalho
            novo_cabecalho = df_resumo.iloc[0].astype(str)
            df_resumo.columns = novo_cabecalho
            # Remover a primeira linha, que agora é o cabeçalho duplicado
            df_resumo = df_resumo.iloc[1:].reset_index(drop=True)
            
            '''--calculo nescessario resumo--'''
            def encontrar_VALOR_positivo_resumo(coluna_filtro, linha_descricao, coluna_a_encontrar,contem=True):
                try:
                    # Tratar valores ausentes substituindo por uma string vazia
                    df_resumo[coluna_filtro] = df_resumo[coluna_filtro].fillna('')
                    if contem:
                        #encontrar linha a partir de uma coluna como filtro e contendo os valores
                        linha_contas_resultado = df_resumo[df_resumo[coluna_filtro].str.contains(linha_descricao, case=False)]
                    else:
                        # Localizar a célula que contém o texto inserido
                        linha_contas_resultado = df_resumo[df_resumo[coluna_filtro] == linha_descricao]
                    # Obter o valor da coluna inserido
                    valor_linha = linha_contas_resultado[coluna_a_encontrar].str.replace(' ','').values[0]
                    return valor_linha
                except Exception as e:
                    return 0

            def encontrar_VALOR_negativo_resumo(coluna_filtro, linha_descricao,coluna_a_encontrar,contem=True):
                try:
                    # Tratar valores ausentes substituindo por uma string vazia
                    df_resumo[coluna_filtro] = df_resumo[coluna_filtro].fillna('')
                    if contem:
                        # Encontrar linha a partir de uma coluna como filtro e contendo os valores
                        linha_contas_resultado = df_resumo[df_resumo[coluna_filtro].str.contains(linha_descricao, case=False)]
                    else:
                        # Localizar a célula que contém o texto inserido
                        linha_contas_resultado = df_resumo[df_resumo[coluna_filtro] == linha_descricao]
                    # Obter o índice da linha
                    indice_linha = linha_contas_resultado.index[0]
                    # Retornar o valor específico da coluna
                    valor_coluna = float(linha_contas_resultado.loc[indice_linha, coluna_a_encontrar].replace('DB', '').strip().replace(".", "").replace(",", "."))
                    # Retornar o índice da linha
                    return valor_coluna * -1
                except Exception as e:
                    return 0
            
            def encontrar_VALOR_index_resumo(coluna_filtro, linha_descricao,coluna_a_encontrar,contem=True):
                try:
                    # Tratar valores ausentes substituindo por uma string vazia
                    df_resumo[coluna_filtro] = df_resumo[coluna_filtro].fillna('')
                    if contem: 
                        # Encontrar linha a partir de uma coluna como filtro e contendo os valores
                        linha_contas_resultado = df_resumo[df_resumo[coluna_filtro].str.contains(linha_descricao, case=False)]
                    else:
                        # Localizar a célula que contém o texto inserido
                        linha_contas_resultado = df_resumo[df_resumo[coluna_filtro] == linha_descricao]
                    # Obter o índice da linha
                    indice_linha = linha_contas_resultado.index[0]
                    # Obter o número da coluna onde o valor foi encontrado
                    num_coluna = df_resumo.columns.get_loc(coluna_a_encontrar)
                    # Retornar o índice da linha
                    return indice_linha, num_coluna 
                except Exception as e:
                    return None
                
            # Dicionário para armazenar os valores encontrados em cada mês de MvtoLíquido    
            resultado_liquido = {}
            margem_contabil = {}
            lucro = {}
            receitas_nao_operacionais = {}
            recolhido = {}
            lucro_contabil = {}
            venda = {}
            ir = {}
            add = {}
            csll = {}
            total = {}

            for mes in meses:
                #definindo o ano e mes para procurar em relação a função mês
                coluna_a_encontrar = f'{mes}/{ano}'
                coluna_descricao = 'DescriçãoConta '
                
                #função para encontrar o valor e a index da celula procurada
                def preencher_dicionario_resumo(linha,dicionario,contem_linha=True):
                    if contem_linha:
                        # Executando a função encontrar_VALOR_positivo e salvando em uma lista para calculo futuros    
                        encontrar_valor_positivo = encontrar_VALOR_positivo_resumo(coluna_descricao, linha, coluna_a_encontrar)
                        # Executando a função encontrar_VALOR_index para a localização da celula
                        encontrar_valor_index = encontrar_VALOR_index_resumo(coluna_descricao, linha,coluna_a_encontrar)
                    else:
                        # Executando a função encontrar_VALOR_positivo e salvando em uma lista para calculo futuros    
                        encontrar_valor_positivo = encontrar_VALOR_positivo_resumo(coluna_descricao, linha, coluna_a_encontrar,False)
                        # Executando a função encontrar_VALOR_index para a localização da celula
                        encontrar_valor_index = encontrar_VALOR_index_resumo(coluna_descricao, linha,coluna_a_encontrar,False)

                    if 'DB' in str(encontrar_valor_positivo):
                        if contem_linha:
                            # Executando a função encontrar_VALOR_negativo
                            encontrar_valor_negativo = encontrar_VALOR_negativo_resumo(coluna_descricao, linha,f'MvtoLíquido {mes}/{ano}')
                        else:
                            # Executando a função encontrar_VALOR_negativo
                            encontrar_valor_negativo = encontrar_VALOR_negativo_resumo(coluna_descricao, linha,f'MvtoLíquido {mes}/{ano}',False)
                        #Trocando a str DB por um float salvando na celula e na lista para calculos futuros
                        df_resumo.iloc[encontrar_valor_index[0], encontrar_valor_index[1]] = encontrar_valor_negativo
                        dicionario[f'MvtoLíquido_{mes.lower()}_{ano}'] = encontrar_valor_negativo
                        dicionario[f'Index_linha_{mes.lower()}_{ano}']  = encontrar_valor_index[0]
                        dicionario[f'Index_coluna_{mes.lower()}_{ano}']  = encontrar_valor_index[1]
                    else:
                        dicionario[f'MvtoLíquido_{mes.lower()}_{ano}'] = encontrar_valor_positivo
                        dicionario[f'Index_linha_{mes.lower()}_{ano}']  = encontrar_valor_index[0]
                        dicionario[f'Index_coluna_{mes.lower()}_{ano}']  = encontrar_valor_index[1]
                        
                preencher_dicionario_resumo('RESULTADO LIQUIDO',resultado_liquido)
                preencher_dicionario_resumo('Margem Contábil',margem_contabil)
                preencher_dicionario_resumo('Lucro',lucro,False)
                preencher_dicionario_resumo('Receitas Não Operacionais',receitas_nao_operacionais)
                preencher_dicionario_resumo('Recolhido',recolhido)
                preencher_dicionario_resumo('Lucro Contábil',lucro_contabil)
                preencher_dicionario_resumo('Venda',venda)
                preencher_dicionario_resumo('IR',ir)
                preencher_dicionario_resumo('Add',add)
                preencher_dicionario_resumo('Csll',csll)
                preencher_dicionario_resumo('Total',total)

            for mes in meses_minusculo:
                def preencher_valor_planilha(linha,calculo):
                    linha_ = int(linha[f'Index_linha_{mes}_{ano}'])
                    coluna_ = int(linha[f'Index_coluna_{mes}_{ano}'])
                    df_resumo.iloc[linha_,coluna_] = calculo

                def meses_soma(data,mes_1, mes_2, mes_3):
                    mes1 = data[f'MvtoLíquido_{mes_1}_{ano}']
                    mes2 = data[f'MvtoLíquido_{mes_2}_{ano}']
                    mes3 = data[f'MvtoLíquido_{mes_3}_{ano}']
                    calculo = mes1 + mes2 + mes3
                    return calculo
                
                #calculo e preenchimento do resultado liquido
                valor_resultado1 = contas_resultado[f'MvtoLíquido_{mes}_{ano}']
                valor_resultado2 = recuperacao_despesa[f'MvtoLíquido_{mes}_{ano}']
                
                if valor_resultado2 != 0:
                    calculo_resultado = valor_resultado1 - valor_resultado2
                    preencher_valor_planilha(resultado_liquido,calculo_resultado)
                else:    
                    preencher_valor_planilha(resultado_liquido,  contas_resultado[f'MvtoLíquido_{mes}_{ano}'])

                #calculo e preenchimento da margem contabil
                valor_contabil1 = receitas_operacaional[f'MvtoLíquido_{mes}_{ano}']
                valor_contabil2 = custo_mercadorias[f'MvtoLíquido_{mes}_{ano}']
                
                if valor_contabil2 != 0:
                    calculo_contabil = valor_contabil1 + valor_contabil2
                    preencher_valor_planilha(margem_contabil,calculo_contabil)
                else:    
                    preencher_valor_planilha(margem_contabil, contas_resultado[f'MvtoLíquido_{mes}_{ano}'])
                
                if mes == 'mar':
                    #calculo e preenchimento da venda
                    calculo_venda_1 = meses_soma(venda_mercadorias,'jan','fev','mar')
                    preencher_valor_planilha(venda, calculo_venda_1)
                    #calculo e preenchimento lucro contabil
                    lucro_c = meses_soma(contas_resultado,'jan','fev','mar')
                    preencher_valor_planilha(lucro_contabil, lucro_c)
                    #calculo e preenchimento Receitas Não Operacionais
                    calculo_receita = meses_soma(recuperacao_despesas,'jan','fev','mar')
                    preencher_valor_planilha(receitas_nao_operacionais, calculo_receita)
                    #calculo e preenchimento lucro
                    reultado_lucro = lucro_c - calculo_receita
                    preencher_valor_planilha(lucro,reultado_lucro)
                    #calculo e preenchimento lucro
                    cont_social = contribuicao_social[f'MvtoLíquido_{mes}_{ano}']
                    calculo = cont_social / 9
                    preencher_valor_planilha(recolhido,calculo * -1)
                    #calculo e preenchimento imposto de renda
                    calculo_IR = lucro_c * 0.15
                    preencher_valor_planilha(ir,calculo_IR)
                    #calculo e preenchimento add
                    calculo_add = (lucro_c - 60000) * 0.1
                    preencher_valor_planilha(add, calculo_add)
                    #calculo e preenchimento csll
                    calculo_csll = lucro_c * 0.9 / 10
                    preencher_valor_planilha(csll, calculo_csll)
                    #calculo e preenchimento total
                    calculo_total = calculo_IR + calculo_add + calculo_csll
                    preencher_valor_planilha(total, calculo_total)

                if mes == 'jun':
                    #calculo e preenchimento da venda
                    calculo_venda_2 = meses_soma(venda_mercadorias,'abr','mai','jun')
                    preencher_valor_planilha(venda, calculo_venda_2)
                    #calculo e preenchimento lucro contabil
                    lucro_c = meses_soma(contas_resultado,'abr','mai','jun')
                    preencher_valor_planilha(lucro_contabil, lucro_c)
                    #calculo e preenchimento Receitas Não Operacionais
                    calculo_receita = meses_soma(recuperacao_despesas,'abr','mai','jun')
                    preencher_valor_planilha(receitas_nao_operacionais, calculo_receita)
                    #calculo e preenchimento lucro
                    reultado_lucro = lucro_c - calculo_receita
                    preencher_valor_planilha(lucro,reultado_lucro)
                    #calculo e preenchimento lucro
                    cont_social = contribuicao_social[f'MvtoLíquido_{mes}_{ano}']
                    calculo = cont_social / 9
                    preencher_valor_planilha(recolhido,calculo * -1)
                    #calculo e preenchimento imposto de renda
                    calculo_IR = lucro_c * 0.15
                    preencher_valor_planilha(ir,calculo_IR)
                    #calculo e preenchimento add
                    calculo_add = (lucro_c - 60000) * 0.1
                    preencher_valor_planilha(add, calculo_add)
                    #calculo e preenchimento csll
                    calculo_csll = lucro_c * 0.9 / 10
                    preencher_valor_planilha(csll, calculo_csll)
                    #calculo e preenchimento total
                    calculo_total = calculo_IR + calculo_add + calculo_csll
                    preencher_valor_planilha(total, calculo_total)

                if mes == 'set':
                    #calculo e preenchimento da venda
                    calculo_venda_3 = meses_soma(venda_mercadorias,'jul','ago','set')
                    preencher_valor_planilha(venda, calculo_venda_3)
                    #calculo e preenchimento lucro contabil
                    lucro_c = meses_soma(contas_resultado,'jul','ago','set')
                    preencher_valor_planilha(lucro_contabil, lucro_c)
                    #calculo e preenchimento Receitas Não Operacionais
                    calculo_receita = meses_soma(recuperacao_despesas,'jul','ago','set')
                    preencher_valor_planilha(receitas_nao_operacionais, calculo_receita)
                    #calculo e preenchimento lucro
                    reultado_lucro = lucro_c - calculo_receita
                    preencher_valor_planilha(lucro,reultado_lucro)
                    #calculo e preenchimento lucro
                    cont_social = contribuicao_social[f'MvtoLíquido_{mes}_{ano}']
                    calculo = cont_social / 9
                    preencher_valor_planilha(recolhido,calculo * -1)
                    #calculo e preenchimento imposto de renda
                    calculo_IR = lucro_c * 0.15
                    preencher_valor_planilha(ir,calculo_IR)
                    #calculo e preenchimento add
                    calculo_add = (lucro_c - 60000) * 0.1
                    preencher_valor_planilha(add, calculo_add)
                    #calculo e preenchimento csll
                    calculo_csll = lucro_c * 0.9 / 10
                    preencher_valor_planilha(csll, calculo_csll)
                    #calculo e preenchimento total
                    calculo_total = calculo_IR + calculo_add + calculo_csll
                    preencher_valor_planilha(total, calculo_total)

                if mes == 'dez':
                    #calculo e preenchimento da venda
                    calculo_venda_4 = meses_soma(venda_mercadorias,'out','nov','dez')
                    preencher_valor_planilha(venda, calculo_venda_4)
                    #calculo e preenchimento lucro contabil
                    lucro_c = meses_soma(contas_resultado,'out','nov','dez')
                    preencher_valor_planilha(lucro_contabil, lucro_c)
                    #calculo e preenchimento Receitas Não Operacionais
                    calculo_receita = meses_soma(recuperacao_despesas,'out','nov','dez')
                    preencher_valor_planilha(receitas_nao_operacionais, calculo_receita)
                    #calculo e preenchimento lucro
                    reultado_lucro = lucro_c - calculo_receita
                    preencher_valor_planilha(lucro,reultado_lucro)
                    #calculo e preenchimento lucro
                    cont_social = contribuicao_social[f'MvtoLíquido_{mes}_{ano}']
                    calculo = cont_social / 9
                    preencher_valor_planilha(recolhido,calculo * -1)
                    #calculo e preenchimento imposto de renda
                    calculo_IR = lucro_c * 0.15
                    preencher_valor_planilha(ir,calculo_IR)
                    #calculo e preenchimento add
                    calculo_add = (lucro_c - 60000) * 0.1
                    preencher_valor_planilha(add, calculo_add)
                    #calculo e preenchimento csll
                    calculo_csll = lucro_c * 0.9 / 10
                    preencher_valor_planilha(csll, calculo_csll)
                    #calculo e preenchimento total
                    calculo_total = calculo_IR + calculo_add + calculo_csll
                    preencher_valor_planilha(total, calculo_total)
            
            # Usando o ExcelWriter, cria um arquivo .xlsx, usando engine='xlsxwriter'
            with pd.ExcelWriter('dre_tratada.xlsx', engine='xlsxwriter') as writer:
                df_resumo.to_excel(writer, sheet_name='Resumo', index=False)
                df_sheet1.to_excel(writer, sheet_name='Planilha DRE', index=False)
            
            '''logando a biblioteca openpyxl para tratar o designer do excel 
            --------------------------------------------------------------------------------------------'''
            # Carregar o arquivo Excel tratado
            wb = load_workbook('dre_tratada.xlsx')
            # Especifique o nome da aba que deseja modificar
            nome_da_aba = 'Planilha DRE'
            # Acesse a aba desejada
            aba_existente = wb[nome_da_aba]

            def cor_fundo_celula(numero_index_linha, cor_hex):
                # Especifique a cor de fundo que deseja aplicar #ffffff
                cor_de_fundo = cor_hex 
                # Percorra todas as células da linha especificada e defina a cor de fundo
                for cell in aba_existente[numero_index_linha]:
                    cell.fill = PatternFill(start_color=cor_de_fundo, end_color=cor_de_fundo, fill_type='solid')
                
            def encontrar_indices_texto(linha_descricao):
            # Preencher os valores ausentes na coluna de interesse com um valor padrão
                df_filled = df.fillna({'DescriçãoConta ': ''})
                
                # Localizar as células que contêm o texto inserido
                linhas_contas_resultado = df_filled[df_filled['DescriçãoConta '].str.contains(pat=linha_descricao)]
                return linhas_contas_resultado.index.tolist()

            lista_descrição_azul = (
            'CONTAS DE RESULTADO', 'RECEITAS', 'RECEITAS OPERACIONAL LIQUIDA', 'DEDUCOES DA VENDA DE MERCADORIAS',
            'RECEITAS COMERCIAIS', 'RECEITAS DIVERSAS', 'RECEITAS FINANCEIRAS', 'RECEITAS PRESTACOES SERVICOS',
            'CUSTO DOS SERVICOS PRESTADOS', 'CUSTOS DAS MERCADORIAS VENDIDOS', 'CUSTO DAS MERCADORIAS VENDIDOS - CMV',
            'ESTOQUE INICIAL', 'ENTRADAS DE MERCADORIAS', 'ESTOQUE FINAL', 'CUSTO OPERACIONAL', 'CUSTO DE PRODUCAO',
            'DESPESAS OPERACIONAIS', 'DESPESAS COM PESSOAL', 'PRO LABORE', 'DESPESAS COM PESSOAL DIRETO', 'BENEFICIOS',
            'REMUNERACAO VARIAVEIS', 'SERVICOS TERCEIRIZADOS', 'OUTRAS DESPESAS OPERACIONAIS', 'MANUTENCAO', 'SEGUROS',
            'DESPESAS COM VEICULOS', 'TRANSPORTES E DESLOCAMENTOS', 'VIAGENS E ESTADIAS', 'ENERGIA', 'COMUNICACAO',
            'OUTRAS DESPESAS VENDAS', 'DESPESAS EMBALAGENS', 'DESPESAS CONSUMO INTERNO', 'MANUTENCAO DE IMOVEIS',
            'DESPESAS COM INFORMATICA', 'DESPESAS GERAIS', 'MARKETING', 'DESPESAS DE EXPEDIENTE', 'ALUGUEIS',
            'IMPOSTOS E TAXAS', 'DEPRECIACOES E PROVISOES', 'DESPESAS DIVERSAS', 'HONORARIOS PROFISSIONAIS',
            'DESPESAS FINANCEIRAS', 'SUPERMERCADO AQUINO', 'ASSOCIACOES DE CLASSE', 'DESPESAS FISCAIS',
            'RESULTADO DO EXERCICIO', 'DESPESAS E RECEITAS', 'RECEITAS OPERACIONAIS', 'DESPESAS OPERACIONAIS',
            'CUSTO DA MERCADORIAS VENDIDAS', 'RECEITAS E DESPESAS NAO OPERACIONAIS', 'RESULTADO NA VENDA IMOBILIZADO',
            'PROVISAO DE IMPOSTO S/L', 'PROVISAO DE IMPOSTO S/L', 'TRANSF. PARA RESERVA DE L', 'TRANFERENCIA PARA RESERVA DE L'
        )

            # faz um loop for com a lista para pintar as linhas cuja descrição tem na lista
            for descricao in lista_descrição_azul:
                indices = encontrar_indices_texto(descricao)
                for index in indices:
                    index_int = int(index) + 2  # O deslocamento de 2 é para corresponder às linhas do DataFrame df_resumo
                    cor_fundo_celula(index_int, 'C5D9F1')
        
            cor_fundo_celula(4, 'FFFF00') #amarelo

            # Salvar novamente o arquivo
            wb.save('dre_tratada.xlsx')

        except PermissionError as e:
            print(f'Erro: {e}')
        except Exception as e:
            print(f'Erro: {e}')

if __name__ == "__main__":
    tratamento = programa()
    

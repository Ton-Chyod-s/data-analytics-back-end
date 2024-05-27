import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Color
from tqdm import tqdm

def limp_plan(caminhoPasta,ondeSalvar):
    try:
        inicioContador = time.time()
        nome = 'Resultado Clientes'
        df = pd.read_excel(caminhoPasta)

        #deletar primeira linha
        df.columns = df.loc[0]
        df.drop(0, inplace=True)
        df.reset_index(drop=True, inplace=True)

        # Formatação dos CNPJs
        nome_da_coluna = 'CNPJ'
        if nome_da_coluna in df.columns:
            numero_da_coluna = df.columns.get_loc(nome_da_coluna)
            for indice, valor in enumerate(df['CNPJ']):
                if pd.notnull(valor) and isinstance(valor, int):
                    valor = str(valor)
                    if len(valor) == 14:
                        cnpj_formatado = f"{valor[:2]}.{valor[2:5]}.{valor[5:8]}/{valor[8:12]}-{valor[12:]}"
                    else:
                        valor = "0" + valor
                        cnpj_formatado = f"{valor[:2]}.{valor[2:5]}.{valor[5:8]}/{valor[8:12]}-{valor[12:]}"
                    df.iloc[indice, numero_da_coluna] = cnpj_formatado
                if isinstance(valor, str):
                    pass
                else:
                    df.iloc[indice, numero_da_coluna] = ""

        def clean_value(x):
            if isinstance(x, str):
                x = x.replace('R$ -', '0')
                return x
            elif isinstance(x, (int, float)):
                return x
            else:
                return 0

        df.iloc[:, 6:] = df.iloc[:, 6:].map(clean_value)

        # Salvar DataFrame atualizado no Excel
        df.to_excel(f'{ondeSalvar}{nome}.xlsx', index=False)
        
        # Carregar o arquivo Excel tratado
        wb = load_workbook(f'{ondeSalvar}{nome}.xlsx')
        sheet = wb.active

        font_bold = Font(bold=True)  # Negrito
        # Define a cor de preenchimento (azul marinho)
        fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        # Defina um estilo de borda dupla e azul em todos os lados
        border_style = Border(
            top=Side(style='double', color=Color(rgb='4F81BD')),  # Borda dupla azul no topo
            bottom=Side(style='double', color=Color(rgb='4F81BD')))  # Borda dupla azul na parte inferior

        # Formatar a primeira linha (índice) em negrito e com a cor de preenchimento
        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = font_bold
                cell.fill = fill
                
        # Função para pintar números negativos de vermelho e aplicar negrito
        def format_negative_value(cell):
            if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = Font(bold=True, color="FF0000")  # Negrito e cor vermelha

        # Aplicar formatação às células com números negativos
        for row in sheet.iter_rows(min_row=2, min_col=5, max_col=sheet.max_column):
            for cell in row:
                format_negative_value(cell)

        # Percorra todas as linhas e colunas da planilha
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style

        # Salvar o arquivo Excel com as linhas em negrito
        wb.save(f'{ondeSalvar}{nome}.xlsx')

        finalContador = time.time()
        tempoExecucao = finalContador - inicioContador
        return f'Tempo de execução: <strong>{round(tempoExecucao,2)} seconds</strong>'
    except Exception as e:
        return f'Erro: {e}'

if __name__ == '__main__':
    caminhoPasta = r'\\192.168.1.2\dados\SUPERMERCADO ACOMPANHAMENTO\RESULTADOS CLIENTES - ALEX 2024.xlsx'
    ondeSalvar = r'C:\Users\User\Documents\GitHub\processo-dre-excel\src\ArquivoResCliente\\'

    result = limp_plan(caminhoPasta, ondeSalvar)
    print(result)